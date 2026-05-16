
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import sqlite3
import time
import os
import threading

DB_PATH = "personas.db"
CAMPO_ORDEN = "puntaje_evaluacion"
datos_globales = []


COLOR_FONDO       = "#0F1923"
COLOR_PANEL       = "#162030"
COLOR_TARJETA     = "#1E2D40"
COLOR_BORDE       = "#2A3F58"
COLOR_ACENTO      = "#00C4FF"
COLOR_ACENTO2     = "#00FF9D"
COLOR_ACENTO3     = "#FF6B35"
COLOR_TEXTO       = "#E8F4FD"
COLOR_TEXTO_GRIS  = "#7A9BB5"
COLOR_EXITO       = "#00FF9D"
COLOR_ERROR       = "#FF4757"
COLOR_BURBUJA     = "#FF6B35"
COLOR_INSERCION   = "#00C4FF"
COLOR_SELECCION   = "#A855F7"

FUENTE_TITULO  = ("Consolas", 20, "bold")
FUENTE_SUB     = ("Consolas", 12, "bold")
FUENTE_NORMAL  = ("Consolas", 10)
FUENTE_GRANDE  = ("Consolas", 14, "bold")
FUENTE_PEQUENA = ("Consolas", 9)



def bubble_sort(datos):
    n = len(datos)
    for i in range(n - 1):
        hubo_intercambio = False
        for j in range(n - 1 - i):
            if datos[j][CAMPO_ORDEN] > datos[j + 1][CAMPO_ORDEN]:
                datos[j], datos[j + 1] = datos[j + 1], datos[j]
                hubo_intercambio = True
        if not hubo_intercambio:
            break
    return datos


def insertion_sort(datos):
    for i in range(1, len(datos)):
        elemento_actual = datos[i]
        j = i - 1
        while j >= 0 and datos[j][CAMPO_ORDEN] > elemento_actual[CAMPO_ORDEN]:
            datos[j + 1] = datos[j]
            j -= 1
        datos[j + 1] = elemento_actual
    return datos


def selection_sort(datos):
    n = len(datos)
    for i in range(n - 1):
        indice_minimo = i
        for j in range(i + 1, n):
            if datos[j][CAMPO_ORDEN] < datos[indice_minimo][CAMPO_ORDEN]:
                indice_minimo = j
        if indice_minimo != i:
            datos[i], datos[indice_minimo] = datos[indice_minimo], datos[i]
    return datos


COMPLEJIDADES = {
    "Bubble Sort": {
        "peor":  ("O(n²)", "Arreglo en orden inverso"),
        "medio": ("Θ(n²)", "Arreglo aleatorio"),
        "mejor": ("Ω(n)",  "Arreglo ya ordenado"),
        "color": COLOR_BURBUJA,
    },
    "Insertion Sort": {
        "peor":  ("O(n²)", "Arreglo en orden inverso"),
        "medio": ("Θ(n²)", "Arreglo aleatorio"),
        "mejor": ("Ω(n)",  "Arreglo ya ordenado"),
        "color": COLOR_INSERCION,
    },
    "Selection Sort": {
        "peor":  ("O(n²)", "Siempre recorre todo"),
        "medio": ("Θ(n²)", "Siempre igual"),
        "mejor": ("Ω(n²)", "Sin mejora posible"),
        "color": COLOR_SELECCION,
    },
}



class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Taller – Algoritmos de Ordenamiento")
        self.configure(bg=COLOR_FONDO)
        self.geometry("1000x720")
        self.minsize(900, 650)
        self.resizable(True, True)

        self.tiempos = {}
        self._construir_ui()

    
    def _construir_ui(self):
        
        frame_header = tk.Frame(self, bg=COLOR_PANEL,
                                highlightbackground=COLOR_BORDE,
                                highlightthickness=1)
        frame_header.pack(fill="x", padx=0, pady=0)

        tk.Label(
            frame_header,
            text="◈  ANÁLISIS ASINTÓTICO DE ALGORITMOS",
            font=FUENTE_TITULO, fg=COLOR_ACENTO, bg=COLOR_PANEL,
            pady=14
        ).pack(side="left", padx=24)

        self.lbl_estado = tk.Label(
            frame_header, text="⬤  Sin datos",
            font=FUENTE_NORMAL, fg=COLOR_ERROR, bg=COLOR_PANEL
        )
        self.lbl_estado.pack(side="right", padx=24)

        
        frame_main = tk.Frame(self, bg=COLOR_FONDO)
        frame_main.pack(fill="both", expand=True, padx=16, pady=12)

        
        frame_izq = tk.Frame(frame_main, bg=COLOR_PANEL,
                             highlightbackground=COLOR_BORDE,
                             highlightthickness=1, width=260)
        frame_izq.pack(side="left", fill="y", padx=(0, 10))
        frame_izq.pack_propagate(False)

        tk.Label(
            frame_izq, text="MENÚ PRINCIPAL",
            font=FUENTE_SUB, fg=COLOR_TEXTO_GRIS, bg=COLOR_PANEL,
            pady=10
        ).pack(pady=(16, 4))

        self._separador(frame_izq)

        
        self._boton(
            frame_izq,
            texto="[ 1 ]  CARGAR DATOS",
            color=COLOR_ACENTO2,
            comando=self._cargar_datos
        )

        self._separador(frame_izq)
        tk.Label(
            frame_izq, text="ALGORITMOS",
            font=FUENTE_PEQUENA, fg=COLOR_TEXTO_GRIS, bg=COLOR_PANEL
        ).pack(pady=(8, 2))

        self._boton(frame_izq, "[ 2 ]  BUBBLE SORT",
                    COLOR_BURBUJA,
                    lambda: self._ordenar("Bubble Sort", bubble_sort))
        self._boton(frame_izq, "[ 3 ]  INSERTION SORT",
                    COLOR_INSERCION,
                    lambda: self._ordenar("Insertion Sort", insertion_sort))
        self._boton(frame_izq, "[ 4 ]  SELECTION SORT",
                    COLOR_SELECCION,
                    lambda: self._ordenar("Selection Sort", selection_sort))

        self._separador(frame_izq)

        self._boton(frame_izq, "[ 5 ]  VER COMPARATIVA",
                    COLOR_TEXTO_GRIS,
                    self._ver_comparativa)

        self._boton(frame_izq, "[ 6 ]  SALIR",
                    COLOR_ERROR,
                    self.destroy)

        
        frame_der = tk.Frame(frame_main, bg=COLOR_FONDO)
        frame_der.pack(side="left", fill="both", expand=True)

        
        frame_tarjetas = tk.Frame(frame_der, bg=COLOR_FONDO)
        frame_tarjetas.pack(fill="x", pady=(0, 10))

        self.tarjetas = {}
        for nombre, comp in COMPLEJIDADES.items():
            t = self._tarjeta(frame_tarjetas, nombre, comp)
            self.tarjetas[nombre] = t

        
        tk.Label(
            frame_der, text="CONSOLA DE RESULTADOS",
            font=FUENTE_PEQUENA, fg=COLOR_TEXTO_GRIS, bg=COLOR_FONDO,
            anchor="w"
        ).pack(fill="x")

        self.consola = scrolledtext.ScrolledText(
            frame_der,
            bg="#0A1520", fg=COLOR_TEXTO,
            font=("Consolas", 10),
            bd=0, relief="flat",
            insertbackground=COLOR_ACENTO,
            wrap="word",
            highlightbackground=COLOR_BORDE,
            highlightthickness=1,
            state="disabled"
        )
        self.consola.pack(fill="both", expand=True)

        
        self.consola.tag_config("titulo",  foreground=COLOR_ACENTO,
                                 font=("Consolas", 11, "bold"))
        self.consola.tag_config("exito",   foreground=COLOR_EXITO)
        self.consola.tag_config("error",   foreground=COLOR_ERROR)
        self.consola.tag_config("burbuja", foreground=COLOR_BURBUJA)
        self.consola.tag_config("inserc",  foreground=COLOR_INSERCION)
        self.consola.tag_config("selec",   foreground=COLOR_SELECCION)
        self.consola.tag_config("gris",    foreground=COLOR_TEXTO_GRIS)
        self.consola.tag_config("blanco",  foreground=COLOR_TEXTO)

        self._log("◈  Sistema iniciado. Selecciona una opción del menú.\n",
                  "titulo")
        self._log("→  Primero carga los datos con la opción [1]\n", "gris")

        
        self.progreso = ttk.Progressbar(
            frame_der, mode="indeterminate",
            style="custom.Horizontal.TProgressbar"
        )
        estilo = ttk.Style()
        estilo.theme_use("clam")
        estilo.configure("custom.Horizontal.TProgressbar",
                         troughcolor=COLOR_TARJETA,
                         background=COLOR_ACENTO,
                         bordercolor=COLOR_BORDE,
                         lightcolor=COLOR_ACENTO,
                         darkcolor=COLOR_ACENTO)

    
    def _separador(self, parent):
        tk.Frame(parent, bg=COLOR_BORDE, height=1).pack(
            fill="x", padx=12, pady=6
        )

    def _boton(self, parent, texto, color, comando):
        btn = tk.Button(
            parent, text=texto,
            font=("Consolas", 10, "bold"),
            fg=color, bg=COLOR_TARJETA,
            activeforeground=COLOR_FONDO, activebackground=color,
            bd=0, relief="flat",
            cursor="hand2",
            pady=10, padx=8,
            command=comando,
            anchor="w"
        )
        btn.pack(fill="x", padx=12, pady=3)

        def on_enter(e):
            btn.config(bg=color, fg=COLOR_FONDO)
        def on_leave(e):
            btn.config(bg=COLOR_TARJETA, fg=color)

        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    def _tarjeta(self, parent, nombre, comp):
        color = comp["color"]
        frame = tk.Frame(
            parent, bg=COLOR_TARJETA,
            highlightbackground=color,
            highlightthickness=1
        )
        frame.pack(side="left", expand=True, fill="x",
                   padx=4, pady=2)

        tk.Label(
            frame, text=nombre.upper(),
            font=("Consolas", 9, "bold"),
            fg=color, bg=COLOR_TARJETA, pady=4
        ).pack()

        lbl_tiempo = tk.Label(
            frame, text="-- ms",
            font=("Consolas", 18, "bold"),
            fg=COLOR_TEXTO_GRIS, bg=COLOR_TARJETA
        )
        lbl_tiempo.pack()

        filas = [
            ("Peor",  comp["peor"][0]),
            ("Medio", comp["medio"][0]),
            ("Mejor", comp["mejor"][0]),
        ]
        for etiq, nota in filas:
            f = tk.Frame(frame, bg=COLOR_TARJETA)
            f.pack(fill="x", padx=8, pady=1)
            tk.Label(f, text=f"{etiq}:", font=("Consolas", 8),
                     fg=COLOR_TEXTO_GRIS, bg=COLOR_TARJETA,
                     width=6, anchor="w").pack(side="left")
            tk.Label(f, text=nota, font=("Consolas", 8, "bold"),
                     fg=color, bg=COLOR_TARJETA).pack(side="left")

        tk.Frame(frame, bg=COLOR_TARJETA, height=6).pack()

        return lbl_tiempo

    def _log(self, texto, tag="blanco"):
        self.consola.config(state="normal")
        self.consola.insert("end", texto, tag)
        self.consola.see("end")
        self.consola.config(state="disabled")

    def _limpiar_consola(self):
        self.consola.config(state="normal")
        self.consola.delete("1.0", "end")
        self.consola.config(state="disabled")

    
    def _cargar_datos(self):
        global datos_globales

        if not os.path.exists(DB_PATH):
            messagebox.showerror(
                "Error",
                f"No se encontró '{DB_PATH}'.\n\n"
                "Ejecuta primero:\n  python crear_base_datos.py"
            )
            return

        self._limpiar_consola()
        self._log("◈  CARGANDO DATOS\n", "titulo")
        self._log(f"   Fuente : {DB_PATH}\n", "gris")

        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                "SELECT id_persona, nombre, edad, "
                "puntaje_evaluacion FROM personas"
            )
            filas = cur.fetchall()
            conn.close()

            datos_globales = [dict(f) for f in filas]

            self._log(f"   Total  : {len(datos_globales)} registros\n",
                      "exito")
            self._log(f"   Campo  : '{CAMPO_ORDEN}'\n", "gris")
            self._log("\n   Primeros 3 registros:\n", "gris")
            for r in datos_globales[:3]:
                self._log(
                    f"   ID:{r['id_persona']:>4}  "
                    f"{r['nombre']:<25} "
                    f"Puntaje: {r['puntaje_evaluacion']}\n",
                    "blanco"
                )
            self._log("\n✔  Datos cargados correctamente.\n", "exito")
            self._log("   Ahora elige un algoritmo para ordenar.\n", "gris")

            self.lbl_estado.config(
                text=f"⬤  {len(datos_globales)} registros cargados",
                fg=COLOR_EXITO
            )

        except Exception as e:
            self._log(f"\n✖  Error: {e}\n", "error")

    def _ordenar(self, nombre, funcion):
        if not datos_globales:
            messagebox.showwarning(
                "Sin datos",
                "Primero debes cargar los datos.\nUsa la opción [1]."
            )
            return

        
        thread = threading.Thread(
            target=self._ejecutar_ordenamiento,
            args=(nombre, funcion),
            daemon=True
        )
        thread.start()

    def _ejecutar_ordenamiento(self, nombre, funcion):
        tag_color = {
            "Bubble Sort":    "burbuja",
            "Insertion Sort": "inserc",
            "Selection Sort": "selec",
        }.get(nombre, "blanco")

        self._limpiar_consola()
        self._log(f"◈  {nombre.upper()}\n", "titulo")
        self._log(f"   Ordenando {len(datos_globales)} registros...\n",
                  "gris")

        # Mostrar barra de progreso
        self.progreso.pack(fill="x", pady=(4, 0))
        self.progreso.start(10)

        copia = [dict(r) for r in datos_globales]

    
        inicio = time.perf_counter()
        datos_ordenados = funcion(copia)
        fin = time.perf_counter()
        

        tiempo_ms = (fin - inicio) * 1000
        self.tiempos[nombre] = tiempo_ms

        self.progreso.stop()
        self.progreso.pack_forget()

        comp = COMPLEJIDADES[nombre]

        self._log(f"\n   ┌{'─'*46}┐\n", tag_color)
        self._log(f"   │  TIEMPO DE EJECUCIÓN : "
                  f"{tiempo_ms:.4f} ms".ljust(46) + "│\n", tag_color)
        self._log(f"   │  REGISTROS ORDENADOS : "
                  f"{len(datos_ordenados)}".ljust(46) + "│\n", tag_color)
        self._log(f"   └{'─'*46}┘\n\n", tag_color)

        self._log("   COMPLEJIDAD ALGORÍTMICA TEÓRICA\n", "titulo")
        self._log(f"   Peor caso  (Big O)  → {comp['peor'][0]:8s}"
                  f"  {comp['peor'][1]}\n", tag_color)
        self._log(f"   Caso medio (Big Θ)  → {comp['medio'][0]:8s}"
                  f"  {comp['medio'][1]}\n", tag_color)
        self._log(f"   Mejor caso (Big Ω) → {comp['mejor'][0]:8s}"
                  f"  {comp['mejor'][1]}\n", tag_color)

        self._log("\n   VISTA PREVIA (primeros 5 ordenados):\n", "gris")
        for r in datos_ordenados[:5]:
            self._log(
                f"   Puntaje: {r['puntaje_evaluacion']:>6}  │  "
                f"{r['nombre']:<25} │  Edad: {r['edad']}\n",
                "blanco"
            )

        self._log(f"\n✔  {nombre} completado.\n", "exito")

        # Actualizar tarjeta
        self.tarjetas[nombre].config(
            text=f"{tiempo_ms:.2f} ms",
            fg=comp["color"]
        )

        
        self._exportar_excel(datos_ordenados, nombre, tiempo_ms)

    def _exportar_excel(self, datos, nombre, tiempo_ms):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill,
                                          Alignment, Border, Side)
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Ordenados"

            borde = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin")
            )

            ws.merge_cells("A1:D1")
            ws["A1"] = f"Resultado – {nombre}"
            ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", fgColor="2F5496")
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:D2")
            ws["A2"] = (f"Registros: {len(datos)}   |   "
                        f"Tiempo: {tiempo_ms:.4f} ms   |   "
                        f"Campo: {CAMPO_ORDEN}")
            ws["A2"].font = Font(italic=True, size=10, color="FFFFFF")
            ws["A2"].fill = PatternFill("solid", fgColor="2F5496")
            ws["A2"].alignment = Alignment(horizontal="center")

            encabezados = ["ID", "Nombre", "Edad", "Puntaje Evaluacion"]
            for col, t in enumerate(encabezados, 1):
                c = ws.cell(row=3, column=col, value=t)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="1F3864")
                c.alignment = Alignment(horizontal="center")
                c.border = borde

            for i, r in enumerate(datos, 1):
                fila = i + 3
                fill = PatternFill("solid",
                                   fgColor="DCE6F1" if i % 2 == 0
                                   else "FFFFFF")
                for col, val in enumerate(
                    [r["id_persona"], r["nombre"],
                     r["edad"], r["puntaje_evaluacion"]], 1
                ):
                    c = ws.cell(row=fila, column=col, value=val)
                    c.font = Font(size=10)
                    c.fill = fill
                    c.border = borde
                    c.alignment = Alignment(
                        horizontal="left" if col == 2 else "center"
                    )

            for col, w in zip([1, 2, 3, 4], [8, 26, 8, 22]):
                ws.column_dimensions[get_column_letter(col)].width = w

            archivo = (f"resultado_"
                       f"{nombre.lower().replace(' ', '_')}.xlsx")
            wb.save(archivo)
            self._log(f"   Excel guardado: {archivo}\n", "exito")

        except Exception as e:
            self._log(f"   Excel no generado: {e}\n", "error")

    def _ver_comparativa(self):
        if not self.tiempos:
            messagebox.showinfo(
                "Sin datos",
                "Aún no has ejecutado ningún algoritmo."
            )
            return

        self._limpiar_consola()
        self._log("◈  TABLA COMPARATIVA DE TIEMPOS\n\n", "titulo")
        self._log(
            f"   {'#':<4} {'Algoritmo':<22} {'Tiempo (ms)':>12} "
            f"{'Tiempo (s)':>12}\n", "gris"
        )
        self._log(f"   {'─'*52}\n", "gris")

        ordenado = sorted(self.tiempos.items(), key=lambda x: x[1])
        medallas = ["🥇", "🥈", "🥉"]

        for i, (alg, ms) in enumerate(ordenado):
            tag = {
                "Bubble Sort":    "burbuja",
                "Insertion Sort": "inserc",
                "Selection Sort": "selec",
            }.get(alg, "blanco")
            medal = medallas[i] if i < 3 else "  "
            self._log(
                f"   {medal}  {alg:<22} {ms:>10.4f} ms "
                f"  {ms/1000:>10.6f} s\n", tag
            )

        if len(self.tiempos) >= 2:
            mas_rapido = ordenado[0][0]
            mas_lento  = ordenado[-1][0]
            diff = ordenado[-1][1] - ordenado[0][1]
            self._log(f"\n   {'─'*52}\n", "gris")
            self._log(
                f"\n   ✔ Más rápido : {mas_rapido} "
                f"({ordenado[0][1]:.4f} ms)\n", "exito"
            )
            self._log(
                f"   ✖ Más lento  : {mas_lento} "
                f"({ordenado[-1][1]:.4f} ms)\n", "error"
            )
            self._log(
                f"   ↕ Diferencia : {diff:.4f} ms\n", "gris"
            )




if __name__ == "__main__":
    app = App()
    app.mainloop()