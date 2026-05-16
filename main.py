"""
main.py
-------
Taller Práctico: Implementación y Análisis Asintótico
       de Algoritmos de Ordenamiento

Algoritmos implementados desde cero (sin funciones nativas):
  - Bubble Sort    (Ordenamiento Burbuja)
  - Insertion Sort (Ordenamiento Inserción)
  - Selection Sort (Ordenamiento Selección)

Cada ejecución mide el tiempo real de ordenamiento,
muestra la complejidad teórica (Big O, Big Θ, Big Ω)
y exporta el resultado a un archivo Excel.
"""

import sqlite3
import time
import sys
import os

# ─────────────────────────────────────────────────────────────
#  CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────
DB_PATH     = "personas.db"
CAMPO_ORDEN = "puntaje_evaluacion"   # campo por el que se ordena


# ═════════════════════════════════════════════════════════════
#  ALGORITMOS DE ORDENAMIENTO  (implementación pura, sin sort)
# ═════════════════════════════════════════════════════════════

def bubble_sort(datos):
    """
    Bubble Sort – Ordenamiento Burbuja.
    Compara pares adyacentes e intercambia si están en el orden
    incorrecto. Repite hasta que ningún intercambio ocurre.

    Complejidad:
        Peor caso   O(n²)  – arreglo invertido
        Caso medio  Θ(n²)  – arreglo aleatorio
        Mejor caso  Ω(n)   – arreglo ya ordenado (con bandera)
    """
    n = len(datos)
    for i in range(n - 1):
        hubo_intercambio = False
        for j in range(n - 1 - i):
            if datos[j][CAMPO_ORDEN] > datos[j + 1][CAMPO_ORDEN]:
                datos[j], datos[j + 1] = datos[j + 1], datos[j]
                hubo_intercambio = True
        # Optimización: si no hubo intercambios, ya está ordenado
        if not hubo_intercambio:
            break
    return datos


def insertion_sort(datos):
    """
    Insertion Sort – Ordenamiento por Inserción.
    Toma cada elemento y lo inserta en la posición correcta
    dentro de la parte ya ordenada del arreglo.

    Complejidad:
        Peor caso   O(n²)  – arreglo invertido
        Caso medio  Θ(n²)  – arreglo aleatorio
        Mejor caso  Ω(n)   – arreglo ya ordenado
    """
    for i in range(1, len(datos)):
        elemento_actual = datos[i]
        j = i - 1
        while j >= 0 and datos[j][CAMPO_ORDEN] > elemento_actual[CAMPO_ORDEN]:
            datos[j + 1] = datos[j]
            j -= 1
        datos[j + 1] = elemento_actual
    return datos


def selection_sort(datos):
    """
    Selection Sort – Ordenamiento por Selección.
    En cada pasada encuentra el elemento mínimo del subarreglo
    no ordenado y lo coloca en su posición definitiva.

    Complejidad:
        Peor caso   O(n²)  – siempre recorre todo el arreglo
        Caso medio  Θ(n²)  – igual en todos los casos
        Mejor caso  Ω(n²)  – no se beneficia de datos ordenados
    """
    n = len(datos)
    for i in range(n - 1):
        indice_minimo = i
        for j in range(i + 1, n):
            if datos[j][CAMPO_ORDEN] < datos[indice_minimo][CAMPO_ORDEN]:
                indice_minimo = j
        if indice_minimo != i:
            datos[i], datos[indice_minimo] = datos[indice_minimo], datos[i]
    return datos


# ═════════════════════════════════════════════════════════════
#  REPORTE DE COMPLEJIDAD
# ═════════════════════════════════════════════════════════════

COMPLEJIDADES = {
    "Bubble Sort": {
        "peor":   ("O(n²)",  "Arreglo en orden inverso: cada elemento burbujea hasta el final"),
        "medio":  ("Θ(n²)",  "Arreglo aleatorio: en promedio n²/4 comparaciones"),
        "mejor":  ("Ω(n)",   "Arreglo ya ordenado: una sola pasada sin intercambios"),
    },
    "Insertion Sort": {
        "peor":   ("O(n²)",  "Arreglo en orden inverso: cada inserción recorre toda la parte ordenada"),
        "medio":  ("Θ(n²)",  "Arreglo aleatorio: en promedio n²/4 comparaciones y movimientos"),
        "mejor":  ("Ω(n)",   "Arreglo ya ordenado: solo comparaciones, ningún movimiento"),
    },
    "Selection Sort": {
        "peor":   ("O(n²)",  "Siempre recorre el subarreglo restante sin importar el orden"),
        "medio":  ("Θ(n²)",  "Siempre n*(n-1)/2 comparaciones, independiente de los datos"),
        "mejor":  ("Ω(n²)",  "Incluso con datos ordenados realiza todas las comparaciones"),
    },
}


def mostrar_complejidad(nombre_algoritmo, tiempo_ms):
    comp = COMPLEJIDADES[nombre_algoritmo]
    separador = "─" * 58
    print(f"\n{separador}")
    print(f"  REPORTE DE COMPLEJIDAD – {nombre_algoritmo}")
    print(separador)
    print(f"  Tiempo de ejecución : {tiempo_ms:.4f} ms  "
          f"({tiempo_ms / 1000:.6f} s)")
    print(f"  Registros ordenados : {len(datos_globales)}")
    print(f"  Campo de orden      : '{CAMPO_ORDEN}'")
    print(separador)
    print(f"  Peor caso  (Big O)  : {comp['peor'][0]:8s}  "
          f"→ {comp['peor'][1]}")
    print(f"  Caso medio (Big Θ)  : {comp['medio'][0]:8s}  "
          f"→ {comp['medio'][1]}")
    print(f"  Mejor caso (Big Ω)  : {comp['mejor'][0]:8s}  "
          f"→ {comp['mejor'][1]}")
    print(separador)


# ═════════════════════════════════════════════════════════════
#  CARGA DE DATOS
# ═════════════════════════════════════════════════════════════

datos_globales = []   # lista de diccionarios con los registros


def cargar_datos():
    global datos_globales

    if not os.path.exists(DB_PATH):
        print(f"\n  [ERROR] No se encontró '{DB_PATH}'.")
        print("  Ejecuta primero: python crear_base_datos.py\n")
        return

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()
    cur.execute("SELECT id_persona, nombre, edad, puntaje_evaluacion FROM personas")
    filas = cur.fetchall()
    conn.close()

    # Convertir a lista de diccionarios para facilitar la manipulación
    datos_globales = [dict(fila) for fila in filas]

    print(f"\n  [OK] {len(datos_globales)} registros cargados desde '{DB_PATH}'.")
    print(f"  Primer registro : {datos_globales[0]}")
    print(f"  Último registro : {datos_globales[-1]}")


# ═════════════════════════════════════════════════════════════
#  EXPORTAR A EXCEL
# ═════════════════════════════════════════════════════════════

def exportar_excel(datos_ordenados, nombre_algoritmo, tiempo_ms):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [AVISO] openpyxl no disponible. Omitiendo exportación Excel.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Ordenados"

    # ── Paleta de colores ──────────────────────────────────────
    COLOR_CABECERA    = "1F3864"   # azul oscuro
    COLOR_FILA_PAR    = "DCE6F1"   # azul claro
    COLOR_FILA_IMPAR  = "FFFFFF"   # blanco
    COLOR_RESUMEN     = "2F5496"   # azul medio

    borde_delgado = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin")
    )

    # ── Fila de título ─────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Resultado – {nombre_algoritmo}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14,
                               color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOR_RESUMEN)
    ws["A1"].alignment = Alignment(horizontal="center",
                                   vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Fila de metadatos ──────────────────────────────────────
    ws.merge_cells("A2:D2")
    ws["A2"] = (f"Registros: {len(datos_ordenados)}   |   "
                f"Tiempo: {tiempo_ms:.4f} ms   |   "
                f"Ordenado por: {CAMPO_ORDEN}")
    ws["A2"].font      = Font(name="Arial", italic=True, size=10,
                               color="FFFFFF")
    ws["A2"].fill      = PatternFill("solid", fgColor=COLOR_RESUMEN)
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Encabezados ────────────────────────────────────────────
    encabezados = ["ID", "Nombre", "Edad", "Puntaje Evaluación"]
    for col, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=3, column=col, value=titulo)
        celda.font      = Font(name="Arial", bold=True, color="FFFFFF",
                                size=11)
        celda.fill      = PatternFill("solid", fgColor=COLOR_CABECERA)
        celda.alignment = Alignment(horizontal="center",
                                    vertical="center")
        celda.border    = borde_delgado
    ws.row_dimensions[3].height = 20

    # ── Datos ──────────────────────────────────────────────────
    for i, registro in enumerate(datos_ordenados, start=1):
        fila_excel = i + 3
        color = COLOR_FILA_PAR if i % 2 == 0 else COLOR_FILA_IMPAR
        relleno = PatternFill("solid", fgColor=color)

        valores = [
            registro["id_persona"],
            registro["nombre"],
            registro["edad"],
            registro["puntaje_evaluacion"],
        ]
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_excel, column=col, value=valor)
            celda.font      = Font(name="Arial", size=10)
            celda.fill      = relleno
            celda.border    = borde_delgado
            celda.alignment = Alignment(
                horizontal="center" if col != 2 else "left"
            )

    # ── Ancho de columnas ──────────────────────────────────────
    anchos = [10, 28, 10, 22]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(col)].width = ancho

    # ── Segunda hoja: complejidad ──────────────────────────────
    ws2 = wb.create_sheet("Complejidad Algoritmica")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 50

    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"Complejidad Algorítmica – {nombre_algoritmo}"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13,
                                color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", fgColor=COLOR_RESUMEN)
    ws2["A1"].alignment = Alignment(horizontal="center",
                                    vertical="center")
    ws2.row_dimensions[1].height = 26

    encab2 = ["Caso", "Notación", "Descripción"]
    for col, t in enumerate(encab2, start=1):
        c = ws2.cell(row=2, column=col, value=t)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=COLOR_CABECERA)
        c.alignment = Alignment(horizontal="center")
        c.border    = borde_delgado

    comp = COMPLEJIDADES[nombre_algoritmo]
    filas_comp = [
        ("Peor caso  (Big O)",  comp["peor"][0],  comp["peor"][1]),
        ("Caso medio (Big Θ)",  comp["medio"][0], comp["medio"][1]),
        ("Mejor caso (Big Ω)", comp["mejor"][0], comp["mejor"][1]),
    ]
    colores_comp = ["F4CCCC", "FCE5CD", "D9EAD3"]
    for i, (caso, notacion, desc) in enumerate(filas_comp, start=3):
        relleno = PatternFill("solid", fgColor=colores_comp[i - 3])
        for col, val in enumerate([caso, notacion, desc], start=1):
            c = ws2.cell(row=i, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = relleno
            c.border    = borde_delgado
            c.alignment = Alignment(
                horizontal="center" if col < 3 else "left",
                wrap_text=True
            )

    # ── Guardar ────────────────────────────────────────────────
    nombre_archivo = (f"resultado_{nombre_algoritmo.lower().replace(' ', '_')}"
                      f"_{int(tiempo_ms)}ms.xlsx")
    wb.save(nombre_archivo)
    print(f"\n  [OK] Archivo Excel guardado: '{nombre_archivo}'")
    return nombre_archivo


# ═════════════════════════════════════════════════════════════
#  EJECUTAR ORDENAMIENTO CON MEDICIÓN DE TIEMPO
# ═════════════════════════════════════════════════════════════

def ejecutar_ordenamiento(nombre_algoritmo, funcion_sort):
    if not datos_globales:
        print("\n  [AVISO] Primero debes cargar los datos (opción 1).\n")
        return

    # Copia profunda para no alterar los datos originales
    copia = [dict(r) for r in datos_globales]

    print(f"\n  Ejecutando {nombre_algoritmo} sobre "
          f"{len(copia)} registros...")

    # ──── MEDICIÓN (solo el algoritmo, no la carga) ──────────
    inicio = time.perf_counter()
    datos_ordenados = funcion_sort(copia)
    fin   = time.perf_counter()
    # ─────────────────────────────────────────────────────────

    tiempo_ms = (fin - inicio) * 1000

    mostrar_complejidad(nombre_algoritmo, tiempo_ms)

    # Exportar a Excel
    archivo = exportar_excel(datos_ordenados, nombre_algoritmo, tiempo_ms)

    # Vista previa
    print(f"\n  Vista previa (primeros 5 ordenados por '{CAMPO_ORDEN}'):")
    for r in datos_ordenados[:5]:
        print(f"    {r['puntaje_evaluacion']:>8}  |  {r['nombre']:<25}"
              f"  |  Edad: {r['edad']}")
    print()

    return tiempo_ms, archivo


# ═════════════════════════════════════════════════════════════
#  MENÚ PRINCIPAL
# ═════════════════════════════════════════════════════════════

def menu():
    tiempos = {}

    while True:
        print("\n" + "═" * 58)
        print("   TALLER: ANÁLISIS ASINTÓTICO DE ALGORITMOS DE ORDEN.")
        print("═" * 58)
        print("  [1] Cargar datos desde Base de Datos (personas.db)")
        print("  [2] Ordenar con Bubble Sort    (Burbuja)")
        print("  [3] Ordenar con Insertion Sort (Inserción)")
        print("  [4] Ordenar con Selection Sort (Selección)")
        if tiempos:
            print("  [5] Ver resumen comparativo de tiempos")
        print("  [6] Salir")
        print("─" * 58)

        opcion = input("  Elige una opción: ").strip()

        if opcion == "1":
            cargar_datos()

        elif opcion == "2":
            resultado = ejecutar_ordenamiento("Bubble Sort", bubble_sort)
            if resultado:
                tiempos["Bubble Sort"] = resultado[0]

        elif opcion == "3":
            resultado = ejecutar_ordenamiento("Insertion Sort", insertion_sort)
            if resultado:
                tiempos["Insertion Sort"] = resultado[0]

        elif opcion == "4":
            resultado = ejecutar_ordenamiento("Selection Sort", selection_sort)
            if resultado:
                tiempos["Selection Sort"] = resultado[0]

        elif opcion == "5" and tiempos:
            print("\n" + "─" * 48)
            print("  RESUMEN COMPARATIVO DE TIEMPOS")
            print("─" * 48)
            ordenado = sorted(tiempos.items(), key=lambda x: x[1])
            for i, (alg, ms) in enumerate(ordenado, start=1):
                print(f"  #{i}  {alg:<20} {ms:>10.4f} ms")
            print("─" * 48)

        elif opcion == "6":
            print("\n  Programa finalizado. ¡Hasta luego!\n")
            sys.exit(0)

        else:
            print("\n  [AVISO] Opción no válida. Intenta de nuevo.")


# ═════════════════════════════════════════════════════════════
#  PUNTO DE ENTRADA
# ═════════════════════════════════════════════════════════════

if __name__ == "__main__":
    menu()